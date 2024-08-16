import openpyxl

from config import excel_conf
from dao import basic_user, date_dao, spec_dao
from util import print_excel_util

min_len = 6


def load_spec():
    file_dict = excel_conf.get_spec_file_conf()
    for file_name in file_dict:
        load_one(file_dict[file_name])


def load_one(file_conf):
    xlsx_file = openpyxl.load_workbook(filename="./source/spec/" + file_conf.file_name,
                                       data_only=True)  # 打开文件

    sheet_name = file_conf.read_sheet_name
    sheet = xlsx_file[sheet_name]  # 读取名为Sheet1的表
    load_sheet(default_sheet=sheet, file_conf=file_conf)


def load_sheet(default_sheet, file_conf):
    row_max = default_sheet.max_row  # 获取最大行
    title_size = 1
    region_column = 0
    name_column = 0
    cost_rmb_column = 0
    cost_month_column = 0

    for i in range(100):
        val = default_sheet.cell(row=1, column=i + 1).value
        if val is not None and str(val).replace('\n', '').replace('\r', '').replace('\t', '') \
                .replace(' ', '') == file_conf.region_row:
            region_column = i + 1
        if val is not None and str(val).replace('\n', '').replace('\r', '').replace('\t', '') \
                .replace(' ', '') == file_conf.name_row:
            name_column = i + 1
        if val is not None and str(val).replace('\n', '').replace('\r', '').replace('\t', '') \
                .replace(' ', '') == file_conf.read_cost_row:
            cost_rmb_column = i + 1
        if val is not None and str(val).replace('\n', '').replace('\r', '').replace('\t', '') \
                .replace(' ', '') == file_conf.read_cost_month_row:
            cost_month_column = i + 1

    if region_column == 0 or name_column == 0 or cost_rmb_column == 0 or cost_month_column == 0:
        print("加载spec文件未找到列.文件名：" + file_conf.file_name)
        return
    for x in range(row_max - title_size):
        region = default_sheet.cell(row=x + title_size + 1, column=region_column).value
        true_name = default_sheet.cell(row=x + title_size + 1, column=name_column).value
        cost_rmb = default_sheet.cell(row=x + title_size + 1, column=cost_rmb_column).value
        cost_month_val = default_sheet.cell(row=x + title_size + 1, column=cost_month_column).value
        cost_month = None
        if cost_month_val is not None and cost_month_val != '':
            if str(cost_month_val).isdigit():
                date_obj = openpyxl.utils.datetime.from_excel(cost_month_val)
                cost_month = str(date_obj.year) + "年" + str(date_obj.month) + "月份"
        if true_name is None or true_name == "":
            continue
        user = basic_user.getUserInfoByNameAndRegion(true_name, region)
        if user is None:
            print(file_conf.file_name + "已有数据未找到用户" + true_name)
            continue
        spec_dao.insert(user.code, file_conf.file_month, cost_rmb, cost_month)


def print_line(sheet, user_list, column_num, point_seq):
    uids = []
    for one_user in user_list:
        uids.append(one_user.code)

    date_list = spec_dao.select_month_range(uids)
    if len(date_list) == 0:
        date_list = excel_conf.default_month
    if len(date_list) < len(excel_conf.default_month):
        date_list = date_dao.append_pre(date_list, len(excel_conf.default_month) - len(date_list))
    row_len = len(date_list) + 1
    column_num = print_header(sheet, row_len, column_num, point_seq, date_list)
    for user_one in user_list:
        column_num = print_person(sheet, user_one, row_len, column_num, date_list)

    return column_num, point_seq + 1


def print_header(sheet, row_len, cur_row, point_seq, date_list):
    tmp_str = "扣款"

    title = str(point_seq) + "） 2023年 业务员名下电池扣款" + tmp_str + "：  "
    cur_row = print_excel_util.print_first_title(sheet=sheet,
                                                 line_title=title,
                                                 row_len=row_len, cur_row=cur_row)
    arr = ["业务员名下电池扣款"]
    cur_row = print_excel_util.print_table_header(sheet=sheet, table_headers=arr, row_len=row_len,
                                                  cur_row=cur_row)

    arr = ["姓名", tmp_str]
    cur_row = print_excel_util.print_table_header(sheet=sheet, table_headers=arr, row_len=row_len,
                                                  cur_row=cur_row)
    arr = [None]
    arr.extend(date_list)
    cur_row = print_excel_util.print_table_header(sheet=sheet, table_headers=arr, row_len=row_len,
                                                  cur_row=cur_row)
    sheet.merge_cells(start_row=cur_row - 2, end_row=cur_row - 1, start_column=1, end_column=1)
    sheet.row_dimensions[cur_row - 1].height = 35
    return cur_row


def print_person(sheet, user_one, row_len, cur_row, date_list):

    spec_data_list = spec_dao.select_by_uid(user_id=user_one.code)
    arr = [user_one.true_name]
    for date_one in date_list:
        val = None
        for p2p_give_data_one in spec_data_list:
            if p2p_give_data_one.month_val == date_one:
                val = p2p_give_data_one.cost_rmb
                break
        arr.append(val)
    for i in range(row_len):
        val = None
        if len(arr) > i:
            val = arr[i]
        print_excel_util.print_person_cell_val(sheet, cur_row, i + 1, val)
    return cur_row + 1


def get_cost_arr(date_list, cost_detail_list):
    arr = []
    for date_one in date_list:
        cell_data = None
        for cost_detail_one in cost_detail_list:
            if cost_detail_one.cost_month == date_one:
                cell_data = cost_detail_one.cost_rmb
                break
        arr.append(cell_data)
    return arr
