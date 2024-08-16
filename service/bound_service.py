import openpyxl
from openpyxl.styles import Alignment

from config import excel_conf
from dao import basic_user, bound_dao
from util import print_excel_util

temp_cache = {}
temp_cache_sheet_key = "temp_sheet"


def load_bonus():
    print("load bonus")

    file_one = excel_conf.get_bound_file_conf()
    xlsx_file = openpyxl.load_workbook(filename="./source/bound/" + file_one.file_name,
                                       data_only=True)  # 打开文件
    default_sheet = xlsx_file[file_one.read_sheet_name]  # 读取名为Sheet1的表
    row_max = default_sheet.max_row  # 获取最大行
    title_size = 3
    print("load bonus excel success")
    temp_cache[temp_cache_sheet_key] = default_sheet

    for x in range(row_max - title_size):
        # 姓名
        true_name_col = print_excel_util.convert_to_number(file_one.true_name_column)
        company_col = print_excel_util.convert_to_number(file_one.company_column)
        work_col = print_excel_util.convert_to_number(file_one.work_column)
        true_name = default_sheet.cell(row=x + title_size + 1, column=true_name_col).value
        company = default_sheet.cell(row=x + title_size + 1, column=company_col).value
        if company is not None:
            company = company.split('\n')[0]

        work = default_sheet.cell(row=x + title_size + 1, column=work_col).value
        if true_name is None or true_name == "":
            continue
        u = basic_user.getUserInfoByName(true_name, company, work)
        if u is None:
            print("奖金表中用户:" + true_name + ".人员表中未找到")
            continue
        start_col = print_excel_util.convert_to_number(file_one.start_column)
        end_col = print_excel_util.convert_to_number(file_one.end_column)
        pos = start_col
        while pos <= end_col:
            cel_val = default_sheet.cell(row=x + title_size + 1, column=pos).value
            bound_dao.insert(u.code, pos, cel_val)
            pos = pos + 1
    print("load bonus success")


def print_header(sheet, row_len, cur_row, point_seq, file_one):
    title = str(point_seq) + "） " + file_one.bound_title
    cur_row = print_excel_util.print_first_title(sheet=sheet,
                                                 line_title=title,
                                                 row_len=row_len, cur_row=cur_row)

    cur_cell = sheet.cell(cur_row, 1, "姓名")
    cur_cell.border = excel_conf.border
    cur_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cur_cell.font = excel_conf.fontStyle

    target_area_str = print_excel_util.convert_to_letter(2) + str(cur_row) + ':' + print_excel_util.convert_to_letter(
        row_len) + str(cur_row + 1)
    print_excel_util.sheet_copy(temp_cache[temp_cache_sheet_key], sheet, excel_conf.get_bound_table_title_area(),
                                target_area_str=target_area_str)

    print_excel_util.row_merge_none(sheet, cur_row, cur_row + 1, 2, row_len)

    sheet.merge_cells(start_row=cur_row, end_row=cur_row + 1, start_column=1, end_column=1)
    # sheet.row_dimensions[cur_row].height = 45
    return cur_row + 2

    pass


def print_person(sheet, user_one, row_len, cur_row):
    bound_data_list = bound_dao.select_bound_by_uid(uid=user_one.code)
    arr = [user_one.true_name]
    for bound_data_one in bound_data_list:
        arr.append(bound_data_one.rmb)
    for i in range(row_len):
        val = None
        if len(arr) > i:
            val = arr[i]
        print_excel_util.print_person_cell_val(sheet, cur_row, i + 1, val)
    return cur_row + 1


def print_line(sheet, user_list, column_num, point_seq):
    file_one = excel_conf.get_bound_file_conf()
    start_col = print_excel_util.convert_to_number(file_one.start_column)
    end_col = print_excel_util.convert_to_number(file_one.end_column)
    row_len = end_col - start_col + 2

    column_num = print_header(sheet, row_len, column_num, point_seq, file_one)
    for user_one in user_list:
        column_num = print_person(sheet, user_one, row_len, column_num)

    return column_num, point_seq + 1
