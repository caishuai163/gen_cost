from openpyxl.styles import Alignment

from config import excel_conf
from util import print_excel_util


def print_cost_other(sheet, user_list, cur_row, point_seq):
    """

    :param sheet:
    :param user_list:
    :param cur_row:
    :param point_seq:
    :return:
    """
    # cost other
    cur_row = print_cost_other_header(sheet, cur_row, point_seq)
    for user_one in user_list:
        cur_row = print_cost_other_person(sheet, cur_row, user_one)
    return cur_row, point_seq + 1


def print_cost_other_header(sheet, cur_row, point_seq):
    arr = [str(point_seq) + "） 其他扣款：   "]
    for i in range(5):
        val = None
        if len(arr) > i:
            val = arr[i]
        cur_cell = sheet.cell(cur_row, i + 1, val)
        # cur_cell.border = border
        cur_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    sheet.row_dimensions[cur_row].height = 30
    sheet.merge_cells(start_row=cur_row, end_row=cur_row, start_column=1, end_column=5)

    cur_row = cur_row + 1
    arr = ["姓名", "人均扣款总金额", "扣款月数", "月均扣款金额", excel_conf.get_title_other_cost_from_day()]
    for i in range(5):
        val = None
        if len(arr) > i:
            val = arr[i]
        cur_cell = sheet.cell(cur_row, i + 1, val)
        cur_cell.border = excel_conf.border
        cur_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cur_cell.font = excel_conf.fontStyle
    sheet.row_dimensions[cur_row].height = 35
    return cur_row + 1


def print_cost_other_person(sheet, cur_row, user_one):
    """
    暂未实现
    :param sheet:
    :param cur_row:
    :param user_one:
    :return:
    """
    arr = [user_one.true_name, None, None, None,
           None]
    for i in range(len(arr)):
        val = None
        if len(arr) > i:
            val = arr[i]
        print_excel_util.print_person_cell_val(sheet, cur_row, i + 1, val)
    return cur_row + 1
