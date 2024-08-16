import openpyxl
from openpyxl.styles import Alignment

from config import excel_conf
from dao import basic_user, cost
from util import print_excel_util


def load_cost():
    """
    加载扣款文件
    """
    conf_list = excel_conf.get_cost_file_title_conf()
    for conf in conf_list:
        load_one_cost(conf_list[conf])


def load_one_cost(file_conf):
    xlsx_file = openpyxl.load_workbook(filename="./source/cost/" + file_conf.file_name,
                                       data_only=True)  # 打开文件
    sheet_name = file_conf.read_sheet_name
    default_sheet = xlsx_file[sheet_name]  # 读取名为Sheet1的表
    row_max = default_sheet.max_row  # 获取最大行
    title_size = 3
    month_arr = []
    i = 0
    month_start_col = print_excel_util.convert_to_number(file_conf.month_start)
    while True:
        cell = default_sheet.cell(row=2, column=month_start_col + i)
        start_month = cell.value
        if cell.data_type == 'd':  # 'd' 类型代表日期
            month_arr.append(str(start_month.year) + "年" + str(start_month.month) + "月份")
            i = i + 1
        elif str(start_month).isdigit():
            date_obj = openpyxl.utils.datetime.from_excel(start_month)
            month_arr.append(str(date_obj.year) + "年" + str(date_obj.month) + "月份")
            i = i + 1
        elif str(start_month).find("年") >= 0 and str(start_month).find("月") >= 0:
            index_year = str(start_month).index("年")
            index_month = str(start_month).index("月")
            #
            month_arr.append(str(start_month)[index_year-4:index_year] + "年"
                             + str(start_month)[index_year+1:index_month] + "月份")
            i = i + 1
            # break
        else:
            break
    print(file_conf.file_name + str(month_arr))

    true_name_col = print_excel_util.convert_to_number(file_conf.true_name)
    company_col = print_excel_util.convert_to_number(file_conf.company)
    work_col = print_excel_util.convert_to_number(file_conf.work)

    confirm_lost_col = print_excel_util.convert_to_number(file_conf.confirm_lost)
    find_count_col = print_excel_util.convert_to_number(file_conf.find_count)
    confirm_after_find_col = print_excel_util.convert_to_number(file_conf.confirm_after_find)

    avg_cost_col = print_excel_util.convert_to_number(file_conf.avg_cost)
    cost_month_col = print_excel_util.convert_to_number(file_conf.cost_month)
    month_cost_col = print_excel_util.convert_to_number(file_conf.month_cost)

    for x in range(row_max - title_size):

        # 姓名 B D E
        true_name = default_sheet.cell(row=x + title_size + 1, column=true_name_col).value
        company = default_sheet.cell(row=x + title_size + 1, column=company_col).value
        work = default_sheet.cell(row=x + title_size + 1, column=work_col).value
        if true_name is None or true_name == "":
            continue
        if work == "大区经理":
            company = company.split('\n')[0]
        if true_name == "杨帆":
            print(file_conf.file_name + "杨帆")

        user = basic_user.getUserInfoByName(true_name, company, work)
        if user is None:
            print(file_conf.file_name + "user_tb中未找到用户" + true_name)
            continue

        # "盘点后丢失电池数量       "找回数量	"确定丢失电池数量 所在列 G H I
        confirm_lost = default_sheet.cell(row=x + title_size + 1, column=confirm_lost_col).value
        find_count = default_sheet.cell(row=x + title_size + 1, column=find_count_col).value
        confirm_after_find = default_sheet.cell(row=x + title_size + 1, column=confirm_after_find_col).value
        if confirm_after_find is None or confirm_after_find == "":
            confirm_after_find = confirm_lost

        # 扣款金额	"扣款月数"	"月扣款金额" 所在列 R S T
        avg_cost = default_sheet.cell(row=x + title_size + 1, column=avg_cost_col).value
        cost_month = default_sheet.cell(row=x + title_size + 1, column=cost_month_col).value
        month_cost = default_sheet.cell(row=x + title_size + 1, column=month_cost_col).value

        code = cost.insertCost(user.code, file_conf.file_name, confirm_lost, find_count, confirm_after_find, avg_cost,
                               cost_month,
                               month_cost)

        for j in range(len(month_arr)):
            val = default_sheet.cell(row=x + title_size + 1, column=month_start_col + j).value
            if val is None or val == "":
                continue
            cost.insertCostDetail(code, month_arr[j], val)


def get_region_title(one_user):
    region_company = one_user.dealer_name
    confirm_after_find = 0
    cost_month = 0
    cost_list = cost.select_cost_list_title_data(one_user.code, excel_conf.get_cost_base_file())
    if len(cost_list) == 0:
        print(one_user.true_name + "输出文件时生成第一行大标题找不到数据")
    else:
        cost_item = cost_list[0]
        cost_month = cost_item.cost_month
        if cost_item.confirm_after_find is None:
            print(one_user.true_name + "输出文件时生成第一行大标题找不到丢失后找回数据")
        else:
            confirm_after_find = cost_item.confirm_after_find
    tmp_region = one_user.region
    if one_user.work != '大区经理':
        tmp_region = tmp_region + '-' + one_user.part
    return excel_conf.get_title_template().format(region_company, confirm_after_find,
                                                  cost_month,
                                                  confirm_after_find * 1200,
                                                  excel_conf.get_title_bound_from_day(), tmp_region)


def print_cost_detail(sheet, user_list, row_len, cur_row, date_list, file_one, point_seq):
    # batch 1
    file_one_conf = excel_conf.get_cost_file_title_conf()[file_one.file_name]
    cur_row = print_cost_line_header(sheet, cur_row, row_len, date_list,
                                     str(point_seq) + "） " + file_one_conf.row_title,
                                     file_one_conf.row_table_title)
    for user in user_list:
        cur_row = print_cost_line_person(sheet, cur_row, row_len, date_list, user, file_one)
    return cur_row, point_seq + 1


def print_cost(sheet, user_list, column_num, point_seq):
    uids = []
    for one_user in user_list:
        uids.append(one_user.code)
    region_title = get_region_title(user_list[0])
    date_list = cost.select_month_range(uids)
    row_len = len(date_list) + 7

    column_num = print_cost_first(sheet, region_title, row_len, column_num)

    file_list = excel_conf.get_cost_file_title_conf()
    for file_name in file_list:
        column_num, point_seq = print_cost_detail(sheet=sheet, user_list=user_list, row_len=row_len, cur_row=column_num,
                                                  date_list=date_list, file_one=file_list[file_name],
                                                  point_seq=point_seq)
    return column_num, point_seq


def print_cost_first(sheet, region_title, row_len, column_num):
    """
    打印第一个大的合并表头
    :param sheet:
    :param region_title:
    :param row_len:
    :param column_num:
    :return:
    """
    cur_cell = sheet.cell(1, 2, region_title)
    cur_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    sheet.row_dimensions[1].height = 200
    sheet.merge_cells(start_row=1, end_row=1, start_column=2, end_column=row_len)
    return column_num + 1


def print_cost_line_header(sheet, cur_row, row_len, date_list, line_title, table_title):
    """
    打印每一个扣款的表头
    :param sheet:
    :param cur_row:
    :param row_len:
    :param date_list:
    :param line_title:
    :param table_title:
    :return:
    """
    cur_row = print_excel_util.print_first_title(sheet=sheet, line_title=line_title, row_len=row_len, cur_row=cur_row)
    cur_row = print_excel_util.print_table_header(sheet=sheet, table_headers=[table_title], row_len=row_len,
                                                  cur_row=cur_row)
    arr = ["姓名", "电池丢失数量", excel_conf.get_table_row_find_day(), "扣减后扣款电池数量", "扣款"]
    cur_row = print_excel_util.print_table_header(sheet=sheet, table_headers=arr, row_len=row_len,
                                                  cur_row=cur_row)
    arr = [None, None, None, None, "人均扣款总金额", "扣款月数", "月均扣款金额"]
    arr.extend(date_list)
    cur_row = print_excel_util.print_table_header(sheet=sheet, table_headers=arr, row_len=row_len,
                                                  cur_row=cur_row)

    sheet.merge_cells(start_row=cur_row - 2, end_row=cur_row - 1, start_column=1, end_column=1)
    sheet.merge_cells(start_row=cur_row - 2, end_row=cur_row - 1, start_column=2, end_column=2)
    sheet.merge_cells(start_row=cur_row - 2, end_row=cur_row - 1, start_column=3, end_column=3)
    sheet.merge_cells(start_row=cur_row - 2, end_row=cur_row - 1, start_column=4, end_column=4)
    sheet.row_dimensions[cur_row - 1].height = 35
    return cur_row


def print_cost_line_person(sheet, cur_row, row_len, date_list, user_one, file_one):
    """
     打印每一个人扣款的明细
    :param sheet:
    :param cur_row:
    :param row_len:
    :param date_list:
    :param user_one:
    :param file_one:
    :return:
    """

    arr = [user_one.true_name]
    cost_one = cost.select_by_user_and_file(user_one.code, file_one.file_name)
    if cost_one is not None:
        arr.extend([cost_one.confirm_lost, cost_one.find_count, cost_one.confirm_after_find,
                    cost_one.avg_cost, cost_one.cost_month, cost_one.month_cost])
        cost_detail_list = cost.select_detail_by_cost_code(cost_one.code)
        detail_arr = get_detail_arr(date_list=date_list, cost_detail_list=cost_detail_list)
        arr.extend(detail_arr)

    for i in range(row_len):
        val = None
        if len(arr) > i:
            val = arr[i]
        print_excel_util.print_person_cell_val(sheet, cur_row, i + 1, val)

    return cur_row + 1


def get_detail_arr(date_list, cost_detail_list):
    arr = []
    for date_one in date_list:
        cell_data = None
        for cost_detail_one in cost_detail_list:
            if cost_detail_one.month_val == date_one:
                cell_data = cost_detail_one.rmb
                break
        arr.append(cell_data)
    return arr
