from openpyxl.styles import numbers

from config import excel_conf
from dao import cost, p2p_dao, spec_dao
from service import cost_service, p2p_service, spec_service
from util import print_excel_util


def print_header(sheet, row_len, column_num, point_seq, date_list):
    title = str(point_seq) + "） 扣款汇总："

    cur_row = print_excel_util.print_first_title(sheet=sheet,
                                                 line_title=title,
                                                 row_len=row_len, cur_row=column_num)
    arr = ["姓名", "扣款类别", "电池丢失数量", "扣款"]
    cur_row = print_excel_util.print_table_header(sheet=sheet, table_headers=arr, row_len=row_len,
                                                  cur_row=cur_row)

    arr = [None, None, None, "人均扣款总金额", "扣款月数", "月均扣款金额"]
    arr.extend(date_list)
    cur_row = print_excel_util.print_table_header(sheet=sheet, table_headers=arr, row_len=row_len,
                                                  cur_row=cur_row)
    sheet.merge_cells(start_row=cur_row - 2, end_row=cur_row - 1, start_column=1, end_column=1)
    sheet.merge_cells(start_row=cur_row - 2, end_row=cur_row - 1, start_column=1, end_column=1)
    sheet.merge_cells(start_row=cur_row - 2, end_row=cur_row - 1, start_column=2, end_column=2)
    sheet.merge_cells(start_row=cur_row - 2, end_row=cur_row - 1, start_column=3, end_column=3)
    sheet.row_dimensions[cur_row - 1].height = 35
    return cur_row


def print_person_cost(sheet, user_one, row_len, cur_row, date_list, file_one):
    """
        打印每一个人扣款的明细
       :param sheet:
       :param cur_row:
       :param row_len:
       :param date_list:
       :param user_one:
       :param file_one:
       :return:
       """
    arr = [user_one.true_name, file_one.sum_row_text]
    cost_one = cost.select_by_user_and_file(user_one.code, file_one.file_name)
    if cost_one is not None:
        arr.extend([cost_one.confirm_lost,
                    cost_one.avg_cost, cost_one.cost_month, cost_one.month_cost])
        cost_detail_list = cost.select_detail_by_cost_code(cost_one.code)
        detail_arr = cost_service.get_detail_arr(date_list=date_list, cost_detail_list=cost_detail_list)
        arr.extend(detail_arr)

    for i in range(row_len):
        val = None
        if len(arr) > i:
            val = arr[i]
        print_excel_util.print_person_cell_val(sheet, cur_row, i + 1, val)
    return cur_row + 1


def print_person_sum(sheet, user_one, row_len, cur_row, date_list, start_pos):
    arr = [user_one.true_name, "扣款合计：", None, None, None, None]
    for i in range(len(date_list)):
        # 求和的开始单元格地址
        start = sheet.cell(row=start_pos, column=i + 7).coordinate
        # 求和的结束单元格地址
        end = sheet.cell(row=cur_row - 1, column=i + 7).coordinate
        arr.append(f'=SUM({start}:{end})')

    for i in range(row_len):
        val = None
        if len(arr) > i:
            val = arr[i]
        cur_cell = print_excel_util.print_person_cell_val(sheet, cur_row, i + 1, val)
        if i >= 6:
            # 单独处理
            cur_cell.number_format = numbers.FORMAT_NUMBER
    sheet.merge_cells(start_row=cur_row, end_row=cur_row, start_column=2, end_column=6)
    sheet.merge_cells(start_row=start_pos, end_row=cur_row, start_column=1, end_column=1)
    return cur_row + 1


def print_person_p2p_cost(sheet, user_one, row_len, cur_row, date_list, file_one):
    arr = [user_one.true_name, file_one.sum_row_text, None, None, None, None]
    p2p_cost_list = p2p_dao.select_cost_by_uid(user_one.code, file_one.file_month)
    detail_arr = p2p_service.get_cost_arr(date_list=date_list, cost_detail_list=p2p_cost_list)
    arr.extend(detail_arr)

    for i in range(row_len):
        val = None
        if len(arr) > i:
            val = arr[i]
        print_excel_util.print_person_cell_val(sheet, cur_row, i + 1, val)

    return cur_row + 1


def print_person_spec_cost(sheet, user_one, row_len, cur_row, date_list, file_one):
    arr = [user_one.true_name, file_one.sum_row_text, None, None, None, None]
    spec_cost_list = spec_dao.select_cost_by_uid(user_one.code, file_one.file_month)
    detail_arr = spec_service.get_cost_arr(date_list=date_list, cost_detail_list=spec_cost_list)
    arr.extend(detail_arr)

    for i in range(row_len):
        val = None
        if len(arr) > i:
            val = arr[i]
        print_excel_util.print_person_cell_val(sheet, cur_row, i + 1, val)

    return cur_row + 1


def print_person(sheet, user_one, row_len, column_num, date_list):
    file_list = excel_conf.get_cost_file_title_conf()
    start_pos = column_num
    for file_name in file_list:
        column_num = print_person_cost(sheet, user_one, row_len, column_num, date_list, file_list[file_name])
    file_list = excel_conf.get_p2p_file_conf()
    for file_name in file_list:
        column_num = print_person_p2p_cost(sheet, user_one, row_len, column_num, date_list, file_list[file_name])

    file_list = excel_conf.get_spec_file_conf()
    for file_name in file_list:
        column_num = print_person_spec_cost(sheet, user_one, row_len, column_num, date_list, file_list[file_name])

    column_num = print_person_sum(sheet, user_one, row_len, column_num, date_list, start_pos)
    return column_num


def print_line(sheet, user_list, column_num, point_seq):
    uids = []
    for one_user in user_list:
        uids.append(one_user.code)

    date_list = cost.select_sum_month_range(uids)
    row_len = len(date_list) + 6

    column_num = print_header(sheet, row_len, column_num, point_seq, date_list)
    for user_one in user_list:
        column_num = print_person(sheet, user_one, row_len, column_num, date_list)

    return column_num, point_seq + 1
