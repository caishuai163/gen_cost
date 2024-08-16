import openpyxl

from config import excel_conf
from dao import basic_user, p2p_dao, date_dao
from util import print_excel_util

min_len = 6


def load_ab():
    file_dict = excel_conf.get_p2p_file_conf()
    for file_name in file_dict:
        print(file_name)
        load_one(file_dict[file_name])


def load_one(file_conf):
    xlsx_file = openpyxl.load_workbook(filename="./source/p2p/" + file_conf.file_name,
                                       data_only=True, read_only=True)  # 打开文件

    sheet_name = file_conf.read_sheet_name
    sheet = xlsx_file[sheet_name]  # 读取名为Sheet1的表
    load_sheet(default_sheet=sheet, file_conf=file_conf)


def load_sheet(default_sheet, file_conf):
    row_max = default_sheet.max_row  # 获取最大行
    title_size = 2
    give_rmb_column = 0
    cost_rmb_column = 0
    region_column = 0
    part_column = 0
    true_name_column = 0
    work_column = 0

    for i in range(100):
        val = default_sheet.cell(row=2, column=i + 1).value
        if val is not None and str(val).replace('\n', '').replace('\r', '').replace('\t', '') \
                .replace(' ', '') == file_conf.read_give_row:
            give_rmb_column = i + 1
        if val is not None and str(val).replace('\n', '').replace('\r', '').replace('\t', '') \
                .replace(' ', '') == file_conf.read_cost_row:
            cost_rmb_column = i + 1
        if val is not None and str(val).replace('\n', '').replace('\r', '').replace('\t', '') \
                .replace(' ', '') == file_conf.read_cost_month_row:
            cost_month_column = i + 1
        # if val is not None and str(val).replace('\n', '').replace('\r', '').replace('\t', '') \
        #         .replace(' ', '') == file_conf.read_cost_month_row:
        #     cost_month_column = i + 1
        # if val is not None and str(val).replace('\n', '').replace('\r', '').replace('\t', '') \
        #         .replace(' ', '') == file_conf.read_cost_month_row:
        #     cost_month_column = i + 1
        # if val is not None and str(val).replace('\n', '').replace('\r', '').replace('\t', '') \
        #         .replace(' ', '') == file_conf.read_cost_month_row:
        #     cost_month_column = i + 1

    if give_rmb_column == 0 or cost_rmb_column == 0 or cost_month_column == 0:
        print("加载p2p文件未找到列.文件名：" + file_conf.file_name)
        return


    last_region_str = None
    last_part_str = None
    for x in range(row_max - title_size):

        region_str = default_sheet.cell(row=x + title_size + 1, column=2).value
        if region_str is not None and region_str != "":
            last_region_str = region_str
        region = last_region_str

        part_str = default_sheet.cell(row=x + title_size + 1, column=4).value
        if part_str is not None and part_str != "":
            last_part_str = part_str
        part = last_part_str

        true_name = default_sheet.cell(row=x + title_size + 1, column=5).value
        work = default_sheet.cell(row=x + title_size + 1, column=6).value
        if work == '业务员':
            work = '业务'


        give_rmb = default_sheet.cell(row=x + title_size + 1, column=give_rmb_column).value
        cost_rmb = default_sheet.cell(row=x + title_size + 1, column=cost_rmb_column).value
        cost_month_val = default_sheet.cell(row=x + title_size + 1, column=cost_month_column).value
        cost_month = None
        if cost_month_val is not None and cost_month_val != '':
            if str(cost_month_val).isdigit():
                date_obj = openpyxl.utils.datetime.from_excel(cost_month_val)
                cost_month = str(date_obj.year) + "年" + str(date_obj.month) + "月份"
        if true_name is None or true_name == "":
            continue
        user = basic_user.getUserInfoByNameAndPart(true_name, region, None, part, work)
        if user is None:
            print(file_conf.file_name + "已有数据未找到用户" + true_name)
            continue
        p2p_dao.insert(user.code, file_conf.file_month, give_rmb, cost_rmb, cost_month)


def print_line(sheet, user_list, is_give, column_num, point_seq):
    uids = []
    for one_user in user_list:
        uids.append(one_user.code)

    date_list = p2p_dao.select_month_range(uids)
    if len(date_list) == 0:
        date_list = excel_conf.default_month
    if len(date_list) < len(excel_conf.default_month):
        date_list = date_dao.append_pre(date_list, len(excel_conf.default_month) - len(date_list))
    row_len = len(date_list) + 1
    column_num = print_header(sheet, row_len, column_num, point_seq, date_list, is_give)
    for user_one in user_list:
        column_num = print_person(sheet, user_one, row_len, column_num, date_list, is_give)

    return column_num, point_seq + 1


def print_header(sheet, row_len, cur_row, point_seq, date_list, is_give):
    tmp_str = "扣款"

    if is_give == 1:
        tmp_str = "奖励"
    title = str(point_seq) + "） 2023年 点对点" + tmp_str + "：  "
    cur_row = print_excel_util.print_first_title(sheet=sheet,
                                                 line_title=title,
                                                 row_len=row_len, cur_row=cur_row)
    arr = ["点对点"]
    cur_row = print_excel_util.print_table_header(sheet=sheet, table_headers=arr, row_len=row_len,
                                                  cur_row=cur_row)

    arr = ["姓名", tmp_str]
    cur_row = print_excel_util.print_table_header(sheet=sheet, table_headers=arr, row_len=row_len,
                                                  cur_row=cur_row)
    arr = [None]
    arr.extend(date_list)
    cur_row = print_excel_util.print_table_header(sheet=sheet, table_headers=arr, row_len=row_len,
                                                  cur_row=cur_row)
    sheet.merge_cells(start_row=cur_row - 2, end_row=cur_row - 1, start_column=1, end_column=1)
    sheet.row_dimensions[cur_row - 1].height = 35
    return cur_row


def print_person(sheet, user_one, row_len, cur_row, date_list, is_give):

    p2p_give_data_list = p2p_dao.select_by_uid(user_id=user_one.code)
    arr = [user_one.true_name]
    for date_one in date_list:
        val = None
        for p2p_give_data_one in p2p_give_data_list:
            if p2p_give_data_one.month_val == date_one:
                val = p2p_give_data_one.give_rmb if is_give else p2p_give_data_one.cost_rmb
                break
        arr.append(val)
    for i in range(row_len):
        val = None
        if len(arr) > i:
            val = arr[i]
        print_excel_util.print_person_cell_val(sheet, cur_row, i + 1, val)
    return cur_row + 1


def get_cost_arr(date_list, cost_detail_list):
    arr = []
    for date_one in date_list:
        cell_data = None
        for cost_detail_one in cost_detail_list:
            if cost_detail_one.cost_month == date_one:
                cell_data = cost_detail_one.cost_rmb
                break
        arr.append(cell_data)
    return arr
