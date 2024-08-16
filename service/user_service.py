import openpyxl

from dao import basic_user


def reload_file():
    basic_user.remove_all()
    xlsx_file = openpyxl.load_workbook(filename="./source/user_tb.xlsx",
                                       data_only=True)  # 打开文件
    default_sheet = xlsx_file['Sheet1']  # 读取名为Sheet1的表
    row_max = default_sheet.max_row  # 获取最大行
    title_size = 1
    for x in range(row_max - title_size):
        user_name = remove(default_sheet.cell(row=x + title_size + 1, column=1).value)
        region =  remove(default_sheet.cell(row=x + title_size + 1, column=2).value)
        region_manager =  remove(default_sheet.cell(row=x + title_size + 1, column=3).value)
        part =  remove(default_sheet.cell(row=x + title_size + 1, column=4).value)
        company =  remove(default_sheet.cell(row=x + title_size + 1, column=5).value)
        work =  remove(default_sheet.cell(row=x + title_size + 1, column=6).value)
        dealer_name =  remove(default_sheet.cell(row=x + title_size + 1, column=7).value)
        platform = default_sheet.cell(row=x + title_size + 1, column=8).value
        need_generated = default_sheet.cell(row=x + title_size + 1, column=9).value

        if user_name is None or user_name == "":
            continue
        basic_user.insert(user_name, region, region_manager, part, company, work, dealer_name, platform, need_generated)


def remove(string_val):
    if string_val is not None:
        if str(string_val).startswith('\n'):
            string_val = str(string_val).removeprefix('\n')
        elif str(string_val).startswith('\r'):
            string_val = str(string_val).removeprefix('\r')
        elif str(string_val).startswith('\t'):
            string_val = str(string_val).removeprefix('\t')
        elif str(string_val).endswith('\n'):
            string_val = str(string_val).removesuffix('\n')
        elif str(string_val).endswith('\r'):
            string_val = str(string_val).removesuffix('\r')
        elif str(string_val).endswith('\t'):
            string_val = str(string_val).removesuffix('\t')
        else:
            return string_val
    return remove(string_val=string_val)

