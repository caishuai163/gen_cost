import openpyxl

from config import excel_conf
from dao import basic_user, ab_dao, date_dao
from util import print_excel_util, excel_reader_util


def load_ab():
    file_dict = excel_conf.get_ab_file_conf()
    for file_name in file_dict:
        load_one(file_name)


def load_one(file):
    xlsx_file = openpyxl.load_workbook(filename="./source/ab/" + file,
                                       data_only=True)  # 打开文件
    file_conf = excel_conf.get_ab_file_conf()[file]

    # pre_sheet_name = excel_conf.get_ab_file_conf()[file].read_pre_sheet_name
    # pre_sheet = xlsx_file[pre_sheet_name]  # 读取名为Sheet1的表
    # load_sheet(default_sheet=pre_sheet, file_conf=file_conf, flag=0)

    after_sheet_name = excel_conf.get_ab_file_conf()[file].read_after_sheet_name
    after_sheet = xlsx_file[after_sheet_name]  # 读取名为Sheet1的表
    load_sheet(default_sheet=after_sheet, file_conf=file_conf, flag=1)


def load_sheet(default_sheet, file_conf, flag):
    row_max = default_sheet.max_row  # 获取最大行
    title_size = file_conf.data_start_row - 1
    region_column = excel_reader_util.get_column(default_sheet, file_conf.read_start_row, file_conf.read_region_row)
    part_column = excel_reader_util.get_column(default_sheet, file_conf.read_start_row, file_conf.read_part_row)
    name_column = excel_reader_util.get_column(default_sheet, file_conf.read_start_row, file_conf.read_name_row)
    work_column = excel_reader_util.get_column(default_sheet, file_conf.read_start_row, file_conf.read_work_row)
    rmb_column = excel_reader_util.get_column(default_sheet, file_conf.read_start_row, file_conf.read_bound_row)

    last_region_str = None
    last_part_str = None
    for x in range(row_max - title_size):

        region_str = default_sheet.cell(row=x + title_size + 1, column=region_column).value
        if region_str is not None and region_str != "":
            last_region_str = region_str

        part_str = default_sheet.cell(row=x + title_size + 1, column=part_column).value
        if part_str is not None and part_str != "":
            last_part_str = part_str
        part = last_part_str

        true_name = default_sheet.cell(row=x + title_size + 1, column=name_column).value
        work = default_sheet.cell(row=x + title_size + 1, column=work_column).value
        if work == '业务员':
            work = '业务'

        rmb = default_sheet.cell(row=x + title_size + 1, column=rmb_column).value
        if true_name is None or true_name == "":
            continue
        # print(last_region_str)
        region = last_region_str.split("-")[0]
        region_manager = last_region_str.split("-")[1]

        user = basic_user.getUserInfoByNameAndPart(true_name, region,region_manager, part, work)
        if user is None:
            print(file_conf.file_name + "已有数据未找到用户" + true_name)
            continue
        ab_dao.insert(user.code, flag, file_conf.file_month, rmb)


def print_line(sheet, user_list, is_after, column_num, point_seq):
    uids = []
    for one_user in user_list:
        uids.append(one_user.code)

    date_list = ab_dao.select_month_range(uids)
    if len(date_list) == 0:
        date_list = excel_conf.default_month
    if len(date_list) < len(excel_conf.default_month):
        date_list = date_dao.append_pre(date_list, len(excel_conf.default_month) - len(date_list))
    row_len = len(date_list) + 1

    column_num = print_header(sheet, row_len, column_num, point_seq, date_list, is_after)
    for user_one in user_list:
        column_num = print_person(sheet, user_one, row_len, column_num, date_list, is_after)

    return column_num, point_seq + 1


def print_header(sheet, row_len, cur_row, point_seq, date_list, is_after):
    title = str(point_seq) + "） 2023年8月31日前 A+B奖励7%：  "
    if is_after == 1:
        title = str(point_seq) + "） 2023年8月31日后 A+B奖励7%：  "

    cur_row = print_excel_util.print_first_title(sheet=sheet,
                                                 line_title=title,
                                                 row_len=row_len, cur_row=cur_row)
    arr = ["A + B"]
    cur_row = print_excel_util.print_table_header(sheet=sheet, table_headers=arr, row_len=row_len,
                                                  cur_row=cur_row)

    arr = ["姓名", "奖励"]
    cur_row = print_excel_util.print_table_header(sheet=sheet, table_headers=arr, row_len=row_len,
                                                  cur_row=cur_row)
    arr = [None]
    arr.extend(date_list)
    cur_row = print_excel_util.print_table_header(sheet=sheet, table_headers=arr, row_len=row_len,
                                                  cur_row=cur_row)
    sheet.merge_cells(start_row=cur_row - 2, end_row=cur_row - 1, start_column=1, end_column=1)
    sheet.row_dimensions[cur_row - 1].height = 35
    return cur_row


def print_person(sheet, user_one, row_len, cur_row, date_list, is_after):

    ab_data_list = ab_dao.select_by_uid(user_id=user_one.code, is_after=is_after)
    arr = [user_one.true_name]
    for date_one in date_list:
        val = None
        for ab_data_one in ab_data_list:
            if ab_data_one.month_val == date_one:
                val = ab_data_one.rmb
                break
        arr.append(val)
    for i in range(row_len):
        val = None
        if len(arr) > i:
            val = arr[i]
        print_excel_util.print_person_cell_val(sheet, cur_row, i + 1, val)

    return cur_row + 1
