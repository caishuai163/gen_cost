import copy
import re

from openpyxl.styles import Alignment, numbers

from config import excel_conf


def print_first_title(sheet, line_title, row_len, cur_row):
    arr = [line_title]
    for i in range(row_len):
        val = None
        if len(arr) > i:
            val = arr[i]
        cur_cell = sheet.cell(cur_row, i + 1, val)
        # cur_cell.border = border
        cur_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    sheet.row_dimensions[cur_row].height = 30  # 设置第2行高度为30
    sheet.merge_cells(start_row=cur_row, end_row=cur_row, start_column=len(arr), end_column=row_len)
    return cur_row + 1


def print_table_header(sheet, table_headers, row_len, cur_row):
    end_column = row_len
    if len(table_headers) > row_len:
        end_column = len(table_headers)

    for i in range(row_len):
        val = None
        if len(table_headers) > i:
            val = table_headers[i]
        cur_cell = sheet.cell(cur_row, i + 1, val)
        cur_cell.border = excel_conf.border
        cur_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cur_cell.font = excel_conf.fontStyle
    sheet.merge_cells(start_row=cur_row, end_row=cur_row, start_column=len(table_headers), end_column=end_column)
    return cur_row + 1


def sheet_copy(ws1, ws2, source_area_str, target_area_str):
    # 分别指定复制和粘贴所在sheet的位置（本文复制粘贴的单元格区域都在ws内，ws是什么在上面已经指定好）
    source_area = ws1[source_area_str]
    target_area = ws2[target_area_str]

    # 创造source_cell_list，用以和target_cell_list一一对应：
    source_cell_list = []
    for source_row in source_area:
        for source_cell in source_row:
            sc_str = str(source_cell)
            point_time = sc_str.count('.')
            sc_str = sc_str.replace('.', '', point_time - 1)
            start = sc_str.find('.')
            sc_str = sc_str[start + 1: -1]
            source_cell_list.append(sc_str)  # 提取出单元格编号的字符串，如'C8'
    # print('source_cell_list:', source_cell_list)
    target_cell_list = []
    for target_row in target_area:
        for target_cell in target_row:
            tc_str = str(target_cell)
            point_time = tc_str.count('.')
            tc_str = tc_str.replace('.', '', point_time - 1)
            start = tc_str.find('.')
            tc_str = tc_str[start + 1: -1]
            target_cell_list.append(tc_str)  # 提取出单元格编号的字符串，如'L10'
    # print('target_cell_list:', target_cell_list)

    # 获取要复制的单元格总个数：
    cells = len(source_cell_list)

    # 提取并复制格式：
    i = 0
    while i <= cells - 1:
        ws2[target_cell_list[0 + i]].data_type = ws1[source_cell_list[0 + i]].data_type
        if ws1[source_cell_list[0 + i]].has_style:
            # ws2[target_cell_list[0 + i]]._style = copy.copy(ws1[source_cell_list[0 + i]]._style)
            ws2[target_cell_list[0 + i]].font = excel_conf.fontStyle
            ws2[target_cell_list[0 + i]].border = excel_conf.border
            # ws2[target_cell_list[0 + i]].fill = copy.copy(ws1[source_cell_list[0 + i]].fill)
            ws2[target_cell_list[0 + i]].number_format = copy.copy(ws1[source_cell_list[0 + i]].number_format)
            # ws2[target_cell_list[0 + i]].protection = copy.copy(ws1[source_cell_list[0 + i]].protection)
            # ws2[target_cell_list[0 + i]].alignment = copy.copy(ws1[source_cell_list[0 + i]].alignment)
        # 通过引用方法粘贴值: ws['']=ws[''].value
        ws2[target_cell_list[0 + i]] = ws1[source_cell_list[0 + i]].value
        i += 1


def row_merge_none(sheet, row_start, row_end, column_start, column_end):
    y = row_start
    while y <= row_end:
        x = column_start
        pos_s = x
        pos_e = x
        while x <= column_end:
            sheet.cell(row=y, column=x).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            if sheet.cell(row=y, column=x).value is None:
                pos_e = x
            else:
                if pos_s < pos_e:
                    sheet.merge_cells(start_row=y, end_row=y, start_column=pos_s,
                                      end_column=pos_e)
                pos_s = x
            x = x + 1
        y = y + 1


def convert_to_number(letter, column_a=1):
    """
    字母列号转数字
    columnA: 你希望A列是第几列(0 or 1)? 默认1
    return: int
    """
    ab = '_ABCDEFGHIJKLMNOPQRSTUVWXYZ'
    letter0 = letter.upper()
    w = 0
    for _ in letter0:
        w *= 26
        w += ab.find(_)
    return w - 1 + column_a


def convert_to_letter(number, column_a=1):
    """
    数字转字母列号
    columnA: 你希望A列是第几列(0 or 1)? 默认1
    return: str in upper case
    """
    ab = '_ABCDEFGHIJKLMNOPQRSTUVWXYZ'
    n = number - column_a
    x = n % 26
    if n >= 26:
        n = int(n / 26)
        return convert_to_letter(n, 1) + ab[x + 1]
    else:
        return ab[x + 1]


def is_number(s):
    pattern = r'^[-+]?[0-9]+(\.[0-9]+)?$'
    return bool(re.match(pattern, str(s)))


def print_person_cell_val(sheet, cur_row, column, val):
    cur_cell = sheet.cell(cur_row, column, val)
    cur_cell.border = excel_conf.border
    cur_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cur_cell.font = excel_conf.fontStyle
    if is_number(val):
        cur_cell.number_format = numbers.FORMAT_NUMBER
        cur_cell.data_type = 'n'
    return cur_cell